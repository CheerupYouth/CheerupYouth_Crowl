# 엑셀 만들기
import openpyxl

# 엑셀파일 만들기
wb = openpyxl.Workbook()

# 엑셀 워크시트 만들기
ws = wb.create_sheet('주거정책')

# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '노승희'

# 엑셀 저장
wb.save(r'C:\Users\tmdgm\Desktop\pyex\주거정책_data.xlsx')
