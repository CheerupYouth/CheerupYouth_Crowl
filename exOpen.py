# 엑셀 불러오기 연습
import openpyxl

fpath = r'C:\Users\tmdgm\Desktop\pyex\주거정책_data.xlsx' # 엑셀 파일 저장된 경로
# 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 엑셀 시트 선택
ws = wb['주거정책']

# 데이터 수정하기
ws['A3'] = 44
ws['B3'] = '김정민'

# 엑셀저장하기
wb.save(fpath)